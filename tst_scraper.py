"""
TST Jurisprudence Scraper
=========================
Coleta decisões monocráticas (e outros tipos) do portal de Pesquisa de
Jurisprudência do Tribunal Superior do Trabalho e exporta para um arquivo Excel.

Como funciona
-------------
A página https://jurisprudencia.tst.jus.br é uma SPA (React) que consome a API
em https://jurisprudencia-backend2.tst.jus.br. O endpoint
    POST /rest/pesquisa-textual/<inicio>/<tamanho>?a=<random>
recebe um JSON com os filtros e devolve os registros já contendo o inteiro teor
no campo `txtConteudoDecisao` — não é preciso fazer uma segunda requisição por
documento.

IMPORTANTE: o hash que aparece na URL (ex.: #53eabf25...) NÃO é interpretado
pelo backend. Os filtros precisam ser informados via parâmetros do dicionário
SEARCH_FILTERS abaixo.

Como executar
-------------
1. Crie um ambiente virtual e instale as dependências:
       python -m venv .venv
       source .venv/bin/activate           # Linux/Mac
       .\.venv\Scripts\activate            # Windows
       pip install requests openpyxl beautifulsoup4

2. Ajuste os parâmetros em SEARCH_FILTERS abaixo para reproduzir a sua
   pesquisa (datas, relator, ementa etc.).

3. Rode:
       python tst_scraper.py

4. O resultado é salvo em decisoes_tst.xlsx no mesmo diretório.
"""

from __future__ import annotations

import argparse
import json
import logging
import random
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração da pesquisa
# ---------------------------------------------------------------------------

# Tipos de jurisprudência (combine os códigos que quiser):
#   "DESPACHO"  -> Decisões Monocráticas (alvo deste exercício)
#   "ACORDAO"   -> Acórdãos
#   "SUM"       -> Súmulas
#   "PN"        -> Precedentes Normativos
#   "OJ"        -> Orientações Jurisprudenciais
#   "DESPGP"    -> Decisões da Presidência
#   "DESPGVP"   -> Decisões da Vice-Presidência
#   "DESPGCG"   -> Decisões da Corregedoria-Geral
SEARCH_FILTERS: dict[str, Any] = {
    # Texto livre (operadores E / OU / NÃO CONTÉM / EXPRESSÃO EXATA)
    "ou": "",
    "e": "",
    "termoExato": "",
    "naoContem": "",
    # Filtros estruturados
    "ementa": " ",
    "dispositivo": "",
    # Número único (deixe os campos vazios para não filtrar)
    "numeracaoUnica": {
        "numero": "", "ano": "", "digito": "",
        "orgao": "5", "tribunal": "", "vara": "",
    },
    # Listas — preencha apenas se você souber os IDs.
    "orgaosJudicantes": [],
    "ministros": [],
    "convocados": [],
    "classesProcessuais": [],
    "indicadores": [],
    "assuntos": [],
    # Qual(is) tipo(s) buscar
    "tipos": ["DESPACHO"],
    # Órgão de origem ("TST" ou "CSJT")
    "orgao": "TST",
    # Datas (formato YYYY-MM-DD ou None)
    "publicacaoInicial": " ",
    "publicacaoFinal": " ",
    "julgamentoInicial": None,
    "julgamentoFinal": None,
    # Ordenação: "data" | "numero" | "relevancia"
    "ordenacao": "data",
}

# Quantos registros por página (a interface usa 100 por padrão).
PAGE_SIZE = 100

# Atraso (em segundos) entre uma página e a próxima para não martelar o servidor.
SLEEP_BETWEEN_PAGES = (0.8, 1.6)  # intervalo aleatório

# Limite máximo de registros a baixar (None = todos).
MAX_RECORDS: int | None = None

# Caminho do arquivo de saída.
OUTPUT_FILE = Path("decisoes_tst.xlsx")

# ---------------------------------------------------------------------------
# Constantes da API
# ---------------------------------------------------------------------------

API_BASE = "https://jurisprudencia-backend2.tst.jus.br"
SEARCH_ENDPOINT = API_BASE + "/rest/pesquisa-textual/{inicio}/{tamanho}"
DOCUMENT_ENDPOINT = API_BASE + "/rest/documentos/{doc_id}"
ORIGIN = "https://jurisprudencia.tst.jus.br"

# Limite máximo de caracteres por célula no Excel (.xlsx) — 32.767.
EXCEL_CELL_LIMIT = 32_000


# ---------------------------------------------------------------------------
# Cliente HTTP
# ---------------------------------------------------------------------------


@dataclass
class TSTClient:
    """Cliente fino para a API de jurisprudência do TST."""

    session: requests.Session = field(default_factory=requests.Session)
    timeout: int = 30

    def __post_init__(self) -> None:
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "pt-BR,pt;q=0.9",
            "Origin": ORIGIN,
            "Referer": ORIGIN + "/",
            "Content-Type": "application/json",
        })

    def search_page(
        self, payload: dict[str, Any], inicio: int, tamanho: int
    ) -> dict[str, Any]:
        """Faz uma requisição POST a /rest/pesquisa-textual/{inicio}/{tamanho}."""
        url = SEARCH_ENDPOINT.format(inicio=inicio, tamanho=tamanho)
        params = {"a": str(random.random())}  # cache-buster usado pelo front

        # Retentativa simples com backoff exponencial.
        for attempt in range(1, 5):
            try:
                resp = self.session.post(
                    url, params=params, data=json.dumps(payload),
                    timeout=self.timeout,
                )
                resp.raise_for_status()
                return resp.json()
            except (requests.RequestException, ValueError) as exc:
                wait = 2 ** attempt
                logging.warning(
                    "Falha ao buscar inicio=%s (%s); nova tentativa em %ss",
                    inicio, exc, wait,
                )
                time.sleep(wait)
        raise RuntimeError(f"Falha definitiva ao buscar página inicio={inicio}")

    def get_document_html(self, doc_id: str) -> str:
        """Baixa o inteiro teor (HTML) avulso, caso necessário."""
        url = DOCUMENT_ENDPOINT.format(doc_id=doc_id)
        resp = self.session.get(url, timeout=self.timeout)
        resp.raise_for_status()
        return resp.text


# ---------------------------------------------------------------------------
# Iterador de registros (paginação)
# ---------------------------------------------------------------------------


def iter_records(
    client: TSTClient,
    filters: dict[str, Any],
    page_size: int = PAGE_SIZE,
    max_records: int | None = None,
) -> Iterator[dict[str, Any]]:
    """Gera registros de todas as páginas, respeitando MAX_RECORDS."""
    inicio = 1
    total = None
    fetched = 0

    while True:
        data = client.search_page(filters, inicio=inicio, tamanho=page_size)
        if total is None:
            total = data.get("totalRegistros", 0)
            logging.info("Total de registros encontrados: %s", total)
            if total == 0:
                return

        registros = data.get("registros", []) or []
        if not registros:
            return

        for envelope in registros:
            yield envelope.get("registro", {})
            fetched += 1
            if max_records and fetched >= max_records:
                logging.info("Limite MAX_RECORDS=%s atingido.", max_records)
                return

        inicio += len(registros)
        if inicio - 1 >= total:
            return

        delay = random.uniform(*SLEEP_BETWEEN_PAGES)
        logging.debug("Aguardando %.2fs antes da próxima página", delay)
        time.sleep(delay)


# ---------------------------------------------------------------------------
# Conversões e normalizações
# ---------------------------------------------------------------------------


def _rtf_to_text(rtf: str) -> str:
    """Conversor mínimo de RTF para texto.

    Em alguns casos o backend devolve o inteiro teor em RTF (ex.: '{\\rtf1...').
    Aqui fazemos uma limpeza simples — para extração robusta o usuário pode
    instalar a biblioteca `striprtf` (pip install striprtf) e adaptar.
    """
    import re
    # Remove grupos de cabeçalho/fonte/cor.
    rtf = re.sub(r"\{\\\*?\\[^{}]+\}", "", rtf)
    # Decodifica caracteres do tipo \'e9 (hex bytes em CP1252).
    def _hex(m: "re.Match[str]") -> str:
        try:
            return bytes.fromhex(m.group(1)).decode("cp1252", "ignore")
        except Exception:
            return ""
    rtf = re.sub(r"\\'([0-9a-fA-F]{2})", _hex, rtf)
    # Remove comandos remanescentes do RTF.
    rtf = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", rtf)
    # Remove chaves.
    rtf = re.sub(r"[{}]", "", rtf)
    # Compacta espaços/linhas.
    rtf = re.sub(r"[ \t]+", " ", rtf)
    lines = [ln.strip() for ln in rtf.splitlines()]
    return "\n".join([ln for ln in lines if ln])


def html_to_text(html: str) -> str:
    """Converte HTML (ou RTF) para texto puro, preservando quebras de linha."""
    if not html:
        return ""
    # Detecta RTF — alguns documentos do TST vêm nesse formato.
    if html.lstrip().startswith("{\\rtf"):
        return _rtf_to_text(html)
    soup = BeautifulSoup(html, "html.parser")
    # Substitui <br> por quebras de linha.
    for br in soup.find_all("br"):
        br.replace_with("\n")
    text = soup.get_text("\n")
    lines = [ln.strip() for ln in text.splitlines()]
    return "\n".join([ln for ln in lines if ln])


def safe_excel_cell(value: str, limit: int = EXCEL_CELL_LIMIT) -> str:
    """Garante que o conteúdo cabe em uma célula do Excel."""
    if value is None:
        return ""
    if len(value) <= limit:
        return value
    return value[: limit - 60] + "\n[... TRUNCADO POR LIMITE DO EXCEL ...]"


def flatten_record(reg: dict[str, Any]) -> dict[str, Any]:
    """Extrai os campos relevantes de um registro retornado pela API."""
    nu = reg.get("numeracaoUnica") or {}
    orgao_julg = reg.get("orgaoJudicante") or {}
    tipo = reg.get("tipo") or {}
    inteiro_html = reg.get("txtConteudoDecisao") or ""
    inteiro_txt = html_to_text(inteiro_html)

    return {
        "id": reg.get("id", ""),
        "tipo": tipo.get("nome", ""),
        "numero_formatado": reg.get("numFormatado", ""),
        "numero_unico": (
            f"{nu.get('numero', '')}-{nu.get('digito', '')}."
            f"{nu.get('ano', '')}.{nu.get('orgao', '')}."
            f"{nu.get('tribunal', '')}.{nu.get('vara', '')}"
            if nu else ""
        ),
        "classe": reg.get("codFase", ""),
        "orgao_julgador": orgao_julg.get("descricao", ""),
        "relator": (reg.get("nomRelator") or "").title(),
        "data_publicacao": reg.get("dtaPublicacao", ""),
        "data_ordenacao": reg.get("dtaOrdenacao", ""),
        "data_atualizacao": reg.get("dtaAtualizacao", ""),
        "tema_proc": reg.get("txtTemaProc", ""),
        "url_inteiro_teor": DOCUMENT_ENDPOINT.format(doc_id=reg.get("id", "")),
        "inteiro_teor_texto": safe_excel_cell(inteiro_txt),
        "inteiro_teor_html": safe_excel_cell(inteiro_html),
    }


# ---------------------------------------------------------------------------
# Saída em Excel
# ---------------------------------------------------------------------------


COLUMN_ORDER = [
    "id", "tipo", "numero_formatado", "numero_unico", "classe",
    "orgao_julgador", "relator",
    "data_publicacao", "data_ordenacao", "data_atualizacao",
    "tema_proc", "url_inteiro_teor",
    "inteiro_teor_texto", "inteiro_teor_html",
]
COLUMN_LABELS = {
    "id": "ID",
    "tipo": "Tipo",
    "numero_formatado": "Processo",
    "numero_unico": "Nº Único",
    "classe": "Classe",
    "orgao_julgador": "Órgão Julgador",
    "relator": "Relator",
    "data_publicacao": "Publicação",
    "data_ordenacao": "Data (ordenação)",
    "data_atualizacao": "Atualização",
    "tema_proc": "Tema",
    "url_inteiro_teor": "URL Inteiro Teor",
    "inteiro_teor_texto": "Inteiro Teor (texto)",
    "inteiro_teor_html": "Inteiro Teor (HTML)",
}


def write_workbook(rows: list[dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisões TST"

    # Cabeçalho com formatação.
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, key in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS[key])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_ORDER))}1"

    # Linhas de dados.
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(COLUMN_ORDER, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))

    # Larguras razoáveis.
    widths = {
        "id": 34, "tipo": 14, "numero_formatado": 32, "numero_unico": 30,
        "classe": 10, "orgao_julgador": 22, "relator": 28,
        "data_publicacao": 14, "data_ordenacao": 14, "data_atualizacao": 22,
        "tema_proc": 14, "url_inteiro_teor": 50,
        "inteiro_teor_texto": 80, "inteiro_teor_html": 80,
    }
    for col_idx, key in enumerate(COLUMN_ORDER, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, 18)

    wb.save(output)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_cli() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Coletor de jurisprudência do TST (decisões monocráticas e outros)."
    )
    parser.add_argument(
        "-o", "--output", default=str(OUTPUT_FILE),
        help="Arquivo de saída .xlsx (padrão: %(default)s)",
    )
    parser.add_argument(
        "-n", "--max-records", type=int, default=MAX_RECORDS,
        help="Limite de registros a baixar (padrão: todos)",
    )
    parser.add_argument(
        "-p", "--page-size", type=int, default=PAGE_SIZE,
        help="Registros por página (padrão: %(default)s)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Logs em nível DEBUG",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_cli()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    client = TSTClient()
    logging.info("Filtros: %s", json.dumps(SEARCH_FILTERS, ensure_ascii=False))

    rows: list[dict[str, Any]] = []
    try:
        for n, reg in enumerate(
            iter_records(client, SEARCH_FILTERS,
                         page_size=args.page_size, max_records=args.max_records),
            start=1,
        ):
            rows.append(flatten_record(reg))
            if n % 50 == 0:
                logging.info("Coletados %s registros...", n)
    except KeyboardInterrupt:
        logging.warning("Interrompido pelo usuário — salvando o que já foi coletado.")

    if not rows:
        logging.warning("Nenhum registro coletado. Verifique seus filtros.")
        return 1

    out = Path(args.output)
    write_workbook(rows, out)
    logging.info("✓ %s registros gravados em %s", len(rows), out.resolve())
    return 0


if __name__ == "__main__":
    sys.exit(main())
